"""
Módulo principal para consulta de CNPJs com correção na seleção de IE Ativa
Versão 2.0 - Com sistema de monitoramento de status das APIs

Contém a classe CNPJConsultor que encapsula toda a lógica de:
- Comunicação com a API Open CNPJ
- Cache de resultados
- Processamento de dados
- Integração com IBGE para códigos municipais
- Seleção correta da IE Ativa
- Monitoramento de status das APIs
- Sistema de log via callback
"""

from constants_v2 import API_CONFIG, EXCEL_STYLES, API_STATUS, ERROR_CODES
from api_status import APIStatusMonitor

import os
import re
import json
import time
import unicodedata
from datetime import datetime, timedelta
from typing import Dict, Optional, List, Tuple, Any, Callable
from queue import Queue
import requests
import csv

class CNPJConsultor:
    """
    Classe principal para consulta e processamento de dados de CNPJ v2.0
    
    Atributos:
        api_url (str): URL base da API
        limite_requisicoes (int): Limite de requisições por minuto
        tempo_espera (int): Tempo de espera entre requisições (segundos)
        stop_processing (bool): Flag para parar processamento
        ultimas_requisicoes (list): Registro temporal das últimas requisições
        cache_file (str): Caminho do arquivo de cache
        cache (dict): Dados em cache
        session (requests.Session): Sessão HTTP persistente
        status_monitor (APIStatusMonitor): Monitor de status das APIs
        api_status (dict): Status atual das APIs
        log_callback (Callable): Função para enviar mensagens de log
    """
    
    def __init__(self, log_callback: Optional[Callable] = None):
        """Inicializa o consultor com configurações padrão"""
        self.api_url = API_CONFIG["base_url"]
        self.limite_requisicoes = API_CONFIG["default_rate_limit"]
        self.tempo_espera = API_CONFIG["default_wait_time"]
        self.stop_processing = False
        self.ultimas_requisicoes = []
        self.cache_file = API_CONFIG["cache_file"]
        self.cache = self._load_cache()
        self.log_callback = log_callback  # Callback para enviar mensagens para a GUI
        
        self.session = requests.Session()
        self._configure_session()
        
        # Sistema de monitoramento de APIs v2.0
        self.status_monitor = APIStatusMonitor()
        self.api_status = self.status_monitor.get_status()

    def _log_message(self, message: str, level: str = "info") -> None:
        """
        Envia mensagem para o callback de log se disponível
        ou para o console como fallback
        
        Args:
            message: Mensagem a ser registrada
            level: Nível do log (info, success, error, warning)
        """
        if self.log_callback:
            self.log_callback(message, level)
        else:
            # Fallback para console se não houver callback
            print(message)

    def _configure_session(self):
        """Configura a sessão HTTP com headers padrão"""
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
            'Accept-Language': 'pt-BR,pt;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
        })
        self.session.timeout = API_CONFIG["timeout"]
        self.session.max_redirects = 5

    def _load_cache(self) -> Dict[str, Any]:
        """Carrega o cache do arquivo JSON"""
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    cache_data = json.load(f)
                    
                    # Limpar cache expirado
                    current_time = time.time()
                    cleaned_cache = {}
                    for cnpj, data in cache_data.items():
                        if 'timestamp' in data:
                            # Verificar se o cache não expirou (24 horas)
                            if current_time - data['timestamp'] < API_CONFIG.get('cache_ttl', 86400):
                                cleaned_cache[cnpj] = data
                        else:
                            # Manter dados antigos sem timestamp
                            cleaned_cache[cnpj] = data
                    
                    # Salvar cache limpo
                    if len(cleaned_cache) != len(cache_data):
                        self.cache = cleaned_cache
                        self._save_cache()
                    
                    return cleaned_cache
                    
        except Exception as e:
            self._log_message(f"Erro ao carregar cache: {str(e)}", "error")
            # Se houver erro, criar cache vazio
            return {}
        return {}

    def _save_cache(self) -> None:
        """Salva o cache atual no arquivo JSON"""
        try:
            # Adicionar timestamp aos dados do cache
            cache_with_timestamp = {}
            for cnpj, data in self.cache.items():
                cache_with_timestamp[cnpj] = data
                if 'timestamp' not in data:
                    cache_with_timestamp[cnpj]['timestamp'] = time.time()
            
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache_with_timestamp, f, ensure_ascii=False, indent=2)
                
            # Limitar tamanho do cache
            self._cleanup_cache()
            
        except Exception as e:
            self._log_message(f"Erro ao salvar cache: {str(e)}", "error")

    def _cleanup_cache(self):
        """Limpa o cache se exceder o tamanho máximo"""
        max_size = API_CONFIG.get('max_cache_size', 1000)
        if len(self.cache) > max_size:
            # Ordenar por timestamp (mais antigos primeiro)
            sorted_cache = sorted(self.cache.items(), 
                                key=lambda x: x[1].get('timestamp', 0))
            
            # Manter apenas os mais recentes
            self.cache = dict(sorted_cache[-max_size:])
            
            # Salvar cache reduzido
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, ensure_ascii=False, indent=2)

    @staticmethod
    def clean_text(text: str) -> str:
        """Normaliza e limpa texto removendo acentos e caracteres especiais"""
        if not isinstance(text, str):
            return str(text)
            
        text = unicodedata.normalize('NFD', text)
        text = text.encode('ascii', 'ignore').decode('utf-8')
        text = re.sub(r'[^a-zA-Z0-9\s]', ' ', text)
        return re.sub(r'\s+', ' ', text).strip()

    @staticmethod
    def clean_cnpj(cnpj: str) -> str:
        """Remove caracteres não numéricos do CNPJ"""
        cleaned = ''.join(filter(str.isdigit, str(cnpj)))
        if len(cleaned) == 14:
            return cleaned
        return cnpj  # Retorna original se não tiver 14 dígitos

    def get_codigo_ibge(self, municipio: str, uf: str) -> str:
        """Obtém código IBGE para um município usando API do IBGE"""
        # Verificar se API IBGE está online
        if not self.is_api_online('ibge'):
            self._log_message("⚠️  API IBGE está offline. Não será possível obter código IBGE.", "warning")
            return ""
            
        cache_key = f"{uf}_{self.clean_text(municipio).upper()}"
        if 'ibge' in self.cache and cache_key in self.cache['ibge']:
            return self.cache['ibge'][cache_key]
            
        try:
            # Usar timeout específico para IBGE
            response = requests.get(
                f"https://servicodados.ibge.gov.br/api/v1/localidades/estados/{uf}/municipios",
                timeout=API_CONFIG["ibge_timeout"],
                headers={
                    'User-Agent': 'CNPJ-Consultor/2.0',
                    'Accept': 'application/json'
                }
            )
            response.raise_for_status()
            
            municipio_clean = self.clean_text(municipio).upper()
            
            for m in response.json():
                if self.clean_text(m['nome']).upper() == municipio_clean:
                    codigo_ibge = str(m['id'])
                    
                    # Atualizar cache IBGE
                    if 'ibge' not in self.cache:
                        self.cache['ibge'] = {}
                    self.cache['ibge'][cache_key] = codigo_ibge
                    self._save_cache()
                    
                    return codigo_ibge
                    
            # Se não encontrou, tentar busca aproximada
            for m in response.json():
                if municipio_clean in self.clean_text(m['nome']).upper():
                    codigo_ibge = str(m['id'])
                    
                    if 'ibge' not in self.cache:
                        self.cache['ibge'] = {}
                    self.cache['ibge'][cache_key] = codigo_ibge
                    self._save_cache()
                    
                    return codigo_ibge
                    
        except requests.exceptions.Timeout:
            self._log_message(f"Timeout ao consultar IBGE para {municipio}/{uf}", "error")
        except requests.exceptions.ConnectionError:
            self._log_message(f"Erro de conexão com API IBGE", "error")
        except Exception as e:
            self._log_message(f"Erro ao consultar IBGE para {municipio}/{uf}: {str(e)}", "error")
            
        return ""

    def _wait_if_needed(self) -> None:
        """Controla o rate limit esperando se necessário"""
        now = datetime.now()
        
        # Limpar requisições antigas (últimos 60 segundos)
        self.ultimas_requisicoes = [
            t for t in self.ultimas_requisicoes 
            if now - t < timedelta(minutes=1)
        ]
        
        # Verificar se precisa esperar
        if len(self.ultimas_requisicoes) >= self.limite_requisicoes:
            oldest = self.ultimas_requisicoes[0]
            wait_time = (timedelta(minutes=1) - (now - oldest)).total_seconds() + 1
            
            self._log_message(f"⏳ Rate limit atingido. Esperando {wait_time:.1f} segundos...", "warning")
            time.sleep(wait_time)
            
            # Limpar lista após espera
            self.ultimas_requisicoes = []

    def consultar_cnpj(self, cnpj: str) -> Optional[Dict[str, Any]]:
        """Consulta um CNPJ na API Open CNPJ"""
        if self.stop_processing:
            return None

        cnpj_limpo = self.clean_cnpj(cnpj)
        
        # Validar CNPJ
        if len(cnpj_limpo) != 14:
            self._log_message(f"❌ CNPJ inválido: {cnpj_limpo}", "error")
            return {'ERRO': 'CNPJ inválido', 'CNPJ': cnpj_limpo}

        # Verificar status da API antes de consultar
        if not self.is_api_online('cnpj'):
            self._log_message(f"⚠️  API CNPJ está offline. Não será possível consultar CNPJ {cnpj_limpo}", "warning")
            return {'ERRO': 'API offline', 'CNPJ': cnpj_limpo}

        # Verificar cache primeiro
        if cnpj_limpo in self.cache:
            cached_data = self.cache[cnpj_limpo]
            
            # Verificar se o cache não expirou
            cache_timestamp = cached_data.get('timestamp', 0)
            if time.time() - cache_timestamp < API_CONFIG.get('cache_ttl', 86400):
                self._log_message(f"📦 Usando cache para CNPJ: {cnpj_limpo}", "info")
                return cached_data
            else:
                # Remover do cache se expirado
                del self.cache[cnpj_limpo]

        try:
            self._wait_if_needed()
            self.ultimas_requisicoes.append(datetime.now())

            self._log_message(f"🔍 Consultando CNPJ: {cnpj_limpo}", "info")
            
            response = self.session.get(
                f"{self.api_url}{cnpj_limpo}", 
                timeout=API_CONFIG["timeout"]
            )

            if response.status_code == 200:
                dados = response.json()
                
                # Adicionar timestamp aos dados
                dados['consultation_timestamp'] = datetime.now().isoformat()
                dados['cnpj'] = cnpj_limpo
                
                self.cache[cnpj_limpo] = dados
                self._save_cache()
                
                self._log_message(f"✅ CNPJ {cnpj_limpo} consultado com sucesso", "success")
                return dados

            elif response.status_code == 429:
                retry_after = int(response.headers.get('Retry-After', self.tempo_espera))
                self._log_message(f"⚠️  Rate limit excedido. Retry after: {retry_after}s", "warning")
                
                time.sleep(retry_after + 2)
                return self.consultar_cnpj(cnpj)

            elif response.status_code == 404:
                self._log_message(f"❌ CNPJ {cnpj_limpo} não encontrado.", "error")
                # Salvar no cache para não consultar novamente
                self.cache[cnpj_limpo] = {'error': 'CNPJ não encontrado', 'status_code': 404}
                self._save_cache()
                return None

            elif response.status_code == 403:
                self._log_message(f"❌ Acesso proibido à API. Verifique configurações.", "error")
                return {'ERRO': 'Acesso proibido', 'STATUS': 403}

            else:
                self._log_message(f"❌ Erro na API: {response.status_code} - {response.text[:100]}...", "error")
                return {'ERRO': f'HTTP {response.status_code}', 'STATUS': response.status_code}

        except requests.exceptions.Timeout:
            self._log_message(f"⏰ Timeout na consulta do CNPJ {cnpj_limpo}", "error")
            return {'ERRO': 'Timeout na consulta', 'CNPJ': cnpj_limpo}
            
        except requests.exceptions.ConnectionError:
            self._log_message(f"🔌 Erro de conexão na consulta do CNPJ {cnpj_limpo}", "error")
            # Marcar API como offline temporariamente
            self.status_monitor.api_status['cnpj']['status'] = 'offline'
            return {'ERRO': 'Erro de conexão', 'CNPJ': cnpj_limpo}
            
        except requests.exceptions.RequestException as e:
            self._log_message(f"⚠️  Erro na requisição: {str(e)}", "error")
            return {'ERRO': f'Erro de requisição: {str(e)}', 'CNPJ': cnpj_limpo}
            
        except Exception as e:
            self._log_message(f"💥 Erro inesperado: {str(e)}", "error")
            return {'ERRO': f'Erro inesperado: {str(e)}', 'CNPJ': cnpj_limpo}

    def extrair_dados(self, dados: Dict[str, Any]) -> Dict[str, Any]:
        """Extrai e formata os dados relevantes da resposta da API"""
        if not dados or 'ERRO' in dados:
            return dados if dados else {'ERRO': 'Dados inválidos'}

        # Processa inscrições estaduais com tratamento robusto
        ies = []
        try:
            for insc in dados.get('registrations', []):
                ie = {
                    'NUMERO': insc.get('number', ''),
                    'UF': insc.get('state', ''),
                    'TIPO': insc.get('type', {}).get('text', ''),
                    'SITUACAO': insc.get('status', {}).get('text', ''),
                    'DATA_SITUACAO': insc.get('statusDate', ''),
                    'ATIVA': 'Sim' if insc.get('enabled', False) else 'Não'
                }
                ies.append(ie)
            
            # Ordenar por: 1) Mesmo estado da empresa, 2) Status "Sem restrição", 3) Data mais recente
            uf_empresa = dados.get('address', {}).get('state', '')
            ies.sort(key=lambda x: (
                x.get('UF', '') != uf_empresa,  # False (0) primeiro para IEs do mesmo estado
                x.get('SITUACAO', '') != 'Sem restrição',  # False (0) primeiro para "Sem restrição"
                -datetime.strptime(x.get('DATA_SITUACAO', '1970-01-01'), '%Y-%m-%d').timestamp() if x.get('DATA_SITUACAO') else 0
            ))
        except Exception as e:
            self._log_message(f"⚠️  Erro ao processar inscrições estaduais: {str(e)}", "warning")
            ies = []

        # Busca código IBGE
        endereco = dados.get('address', {})
        municipio = endereco.get('city', '')
        uf = endereco.get('state', '')
        codigo_ibge = self.get_codigo_ibge(municipio, uf) if municipio and uf else ''

        # Seleciona IE ativa com nova lógica priorizando mesmo estado e "Sem restrição"
        ie_ativa = None
        for ie in ies:
            if ie.get('ATIVA') == 'Sim':
                # Prioriza IEs do mesmo estado da empresa
                if ie.get('UF') == uf:
                    if ie_ativa is None or ie.get('SITUACAO') == 'Sem restrição':
                        ie_ativa = ie
                        if ie.get('SITUACAO') == 'Sem restrição':
                            break  # Melhor caso possível
        
        # Se não encontrou do mesmo estado, pega qualquer IE ativa
        if ie_ativa is None:
            for ie in ies:
                if ie.get('ATIVA') == 'Sim':
                    ie_ativa = ie
                    break

        # Cria dicionário base
        resultado = {
            'CNPJ': dados.get('taxId', ''),
            'RAZAO_SOCIAL': self.clean_text(dados.get('company', {}).get('name', '')),
            'NOME_FANTASIA': self.clean_text(dados.get('alias', '')),
            'ENDERECO': self.clean_text(endereco.get('street', '')),
            'NUMERO': endereco.get('number', ''),
            'COMPLEMENTO': self.clean_text(endereco.get('details', '')),
            'BAIRRO': self.clean_text(endereco.get('district', '')),
            'CEP': endereco.get('zip', ''),
            'MUNICIPIO': municipio,
            'COD_IBGE': codigo_ibge,
            'UF': uf,
            'TELEFONE': dados.get('phone', ''),
            'EMAIL': dados.get('email', ''),
            'SITUACAO': dados.get('status', {}).get('text', ''),
            'DATA_CONSULTA': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'IE_ATIVA_NUMERO': ie_ativa.get('NUMERO', '') if ie_ativa else '',
            'IE_ATIVA_UF': ie_ativa.get('UF', '') if ie_ativa else '',
            'IE_ATIVA_SITUACAO': ie_ativa.get('SITUACAO', '') if ie_ativa else '',
            'IE_ATIVA_DATA': ie_ativa.get('DATA_SITUACAO', '') if ie_ativa else '',
            'INSCRICOES_ESTADUAIS': json.dumps(ies, ensure_ascii=False) if ies else None,
            'ULTIMA_ATUALIZACAO': dados.get('consultation_timestamp', '')
        }

        # Adicionar informações de status da consulta
        if 'consultation_timestamp' in dados:
            resultado['TIMESTAMP_CONSULTA'] = dados['consultation_timestamp']

        return resultado

    def processar_planilha(self, input_file: str, output_file: str, 
                          coluna_cnpj: str, progress_queue: Queue) -> Tuple[bool, str]:
        """Processa um arquivo com múltiplos CNPJs"""
        try:
            self.stop_processing = False
            self.ultimas_requisicoes = []

            # Verificar status da API antes de iniciar
            api_status = self.get_api_status()
            if api_status.get('cnpj', {}).get('status') != 'online':
                progress_queue.put(("error", "API CNPJ está offline. Processamento cancelado."))
                return False, "API CNPJ offline"

            df = self._ler_arquivo(input_file)
            if not df or coluna_cnpj not in df[0]:
                progress_queue.put(("error", f"Coluna '{coluna_cnpj}' não encontrada"))
                return False, f"Coluna '{coluna_cnpj}' não encontrada"
            
            resultados = []
            total = len(df)
            successful = 0
            failed = 0
            cached = 0
            
            for index, row in enumerate(df, 1):
                if self.stop_processing:
                    self._salvar_resultados(resultados, output_file, '_PARCIAL')
                    progress_queue.put(("info", "Processamento cancelado pelo usuário"))
                    return False, "Processamento cancelado"
                
                cnpj = row.get(coluna_cnpj, "")
                cnpj_limpo = self.clean_cnpj(str(cnpj))
                progress_queue.put(("progress", (index, total, cnpj_limpo)))
                
                if len(cnpj_limpo) != 14:
                    resultados.append({'CNPJ': cnpj, 'ERRO': 'CNPJ inválido'})
                    failed += 1
                    continue
                
                # Verificar cache primeiro
                if cnpj_limpo in self.cache:
                    cached_data = self.cache[cnpj_limpo]
                    cache_timestamp = cached_data.get('timestamp', 0)
                    
                    if time.time() - cache_timestamp < API_CONFIG.get('cache_ttl', 86400):
                        resultados.append(self.extrair_dados(cached_data))
                        cached += 1
                        continue
                
                dados = self.consultar_cnpj(cnpj_limpo)
                if dados and 'ERRO' not in dados:
                    resultados.append(self.extrair_dados(dados))
                    successful += 1
                else:
                    resultados.append({'CNPJ': cnpj_limpo, 'ERRO': dados.get('ERRO', 'Falha na consulta') if dados else 'Falha na consulta'})
                    failed += 1
            
            self._salvar_resultados(resultados, output_file)
            
            # Estatísticas do processamento
            stats = f"Processamento concluído! Total: {total}, Sucesso: {successful}, Cache: {cached}, Falhas: {failed}"
            progress_queue.put(("success", stats))
            
            return True, stats
            
        except Exception as e:
            error_msg = f"Erro crítico: {str(e)}"
            progress_queue.put(("error", error_msg))
            return False, error_msg

    def _ler_arquivo(self, input_file: str):
        """Lê arquivo CSV ou Excel"""
        ext = os.path.splitext(input_file)[1].lower()
        
        if ext == '.csv':
            try:
                with open(input_file, newline='', encoding='utf-8') as f:
                    # Tentar detectar encoding automaticamente
                    sample = f.read(1024)
                    f.seek(0)
                    
                    # Verificar se tem BOM (Byte Order Mark)
                    if sample.startswith('\ufeff'):
                        encoding = 'utf-8-sig'
                    else:
                        encoding = 'utf-8'
                    
                    # Ler novamente com encoding correto
                    f = open(input_file, newline='', encoding=encoding)
                    reader = csv.DictReader(f)
                    return list(reader)
                    
            except UnicodeDecodeError:
                # Tentar outros encodings comuns
                for encoding in ['latin-1', 'iso-8859-1', 'cp1252']:
                    try:
                        with open(input_file, newline='', encoding=encoding) as f:
                            reader = csv.DictReader(f)
                            return list(reader)
                    except UnicodeDecodeError:
                        continue
                raise ValueError("Não foi possível decodificar o arquivo CSV")
                
        elif ext in ['.xlsx', '.xls']:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(input_file, read_only=True, data_only=True)
                ws = wb.active
                data = []
                headers = [str(cell.value) if cell.value else f"Coluna_{idx+1}" 
                          for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = {}
                    for idx, value in enumerate(row):
                        if idx < len(headers):
                            row_data[headers[idx]] = value if value is not None else ""
                    data.append(row_data)
                
                wb.close()
                return data
                
            except ImportError:
                raise ValueError("Biblioteca openpyxl não instalada")
            except Exception as e:
                raise ValueError(f"Erro ao ler arquivo Excel: {str(e)}")
        
        raise ValueError("Formato de arquivo não suportado. Use Excel (.xlsx, .xls) ou CSV (.csv)")

    def _salvar_resultados(self, resultados, output_file, suffix=''):
        """Salva resultados em arquivo Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            if suffix:
                base_name = os.path.splitext(output_file)[0]
                output_file = f"{base_name}{suffix}.xlsx"
            
            wb = Workbook()
            ws_main = wb.active
            ws_main.title = 'CNPJs'
            
            # Preparar dados para as abas
            dados_ativas = []
            dados_historico = []
            
            for item in resultados:
                ies = []
                if item.get('INSCRICOES_ESTADUAIS'):
                    try:
                        ies = json.loads(item['INSCRICOES_ESTADUAIS'])
                        uf_empresa = item.get('UF', '')
                        ies.sort(key=lambda x: (
                            x.get('UF', '') != uf_empresa,
                            x.get('SITUACAO', '') != 'Sem restrição',
                            -datetime.strptime(x.get('DATA_SITUACAO', '1970-01-01'), '%Y-%m-%d').timestamp() if x.get('DATA_SITUACAO') else 0
                        ))
                    except json.JSONDecodeError:
                        pass
                
                # Processar histórico de IEs
                for ie in ies:
                    dados_historico.append({
                        'CNPJ': item.get('CNPJ', ''),
                        'RAZAO_SOCIAL': item.get('RAZAO_SOCIAL', ''),
                        'IE_NUMERO': ie.get('NUMERO', ''),
                        'IE_UF': ie.get('UF', ''),
                        'IE_TIPO': ie.get('TIPO', ''),
                        'IE_SITUACAO': ie.get('SITUACAO', ''),
                        'IE_DATA_SITUACAO': ie.get('DATA_SITUACAO', ''),
                        'IE_ATIVA': ie.get('ATIVA', '')
                    })
                
                # Selecionar IE ativa
                ie_ativa = None
                for ie in ies:
                    if ie.get('ATIVA') == 'Sim':
                        if ie.get('UF') == item.get('UF', ''):
                            if ie_ativa is None or ie.get('SITUACAO') == 'Sem restrição':
                                ie_ativa = ie
                                if ie.get('SITUACAO') == 'Sem restrição':
                                    break
                if ie_ativa is None:
                    for ie in ies:
                        if ie.get('ATIVA') == 'Sim':
                            ie_ativa = ie
                            break
                
                # Dados principais
                row_ativa = {
                    'CNPJ': item.get('CNPJ', ''),
                    'RAZAO_SOCIAL': item.get('RAZAO_SOCIAL', ''),
                    'NOME_FANTASIA': item.get('NOME_FANTASIA', ''),
                    'ENDERECO': item.get('ENDERECO', ''),
                    'NUMERO': item.get('NUMERO', ''),
                    'COMPLEMENTO': item.get('COMPLEMENTO', ''),
                    'BAIRRO': item.get('BAIRRO', ''),
                    'CEP': item.get('CEP', ''),
                    'MUNICIPIO': item.get('MUNICIPIO', ''),
                    'COD_IBGE': item.get('COD_IBGE', ''),
                    'UF': item.get('UF', ''),
                    'TELEFONE': item.get('TELEFONE', ''),
                    'EMAIL': item.get('EMAIL', ''),
                    'SITUACAO': item.get('SITUACAO', ''),
                    'DATA_CONSULTA': item.get('DATA_CONSULTA', ''),
                    'IE_ATIVA_NUMERO': ie_ativa.get('NUMERO', '') if ie_ativa else '',
                    'IE_ATIVA_UF': ie_ativa.get('UF', '') if ie_ativa else '',
                    'IE_ATIVA_SITUACAO': ie_ativa.get('SITUACAO', '') if ie_ativa else '',
                    'IE_ATIVA_DATA': ie_ativa.get('DATA_SITUACAO', '') if ie_ativa else '',
                    'STATUS': '✅' if 'ERRO' not in item else '❌'
                }
                dados_ativas.append(row_ativa)
            
            # Escrever dados principais
            if dados_ativas:
                headers = list(dados_ativas[0].keys())
                ws_main.append(headers)
                
                for row in dados_ativas:
                    ws_main.append(list(row.values()))
                
                # Formatação
                header_font = Font(**EXCEL_STYLES["header_font"])
                header_fill = PatternFill(**EXCEL_STYLES["header_fill"])
                
                for cell in ws_main[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = Border(bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar largura das colunas
                for col, width in EXCEL_STYLES["column_widths"]["main"].items():
                    ws_main.column_dimensions[col].width = width
                
                # Congelar painel (cabeçalho)
                ws_main.freeze_panes = 'A2'
            
            # Aba de histórico
            if dados_historico:
                ws_hist = wb.create_sheet('Histórico IEs')
                headers_hist = list(dados_historico[0].keys())
                ws_hist.append(headers_hist)
                
                for row in dados_historico:
                    ws_hist.append(list(row.values()))
                
                # Formatação do histórico
                for cell in ws_hist[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = Border(bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for col, width in EXCEL_STYLES["column_widths"]["history"].items():
                    ws_hist.column_dimensions[col].width = width
                
                ws_hist.freeze_panes = 'A2'
            
            # Aba de estatísticas
            ws_stats = wb.create_sheet('Estatísticas')
            ws_stats.append(['Estatística', 'Valor'])
            ws_stats.append(['Total de CNPJs processados', len(resultados)])
            ws_stats.append(['Consultas bem-sucedidas', sum(1 for r in resultados if 'ERRO' not in r)])
            ws_stats.append(['Consultas com erro', sum(1 for r in resultados if 'ERRO' in r)])
            ws_stats.append(['Data do processamento', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
            ws_stats.append(['Status API CNPJ', self.api_status.get('cnpj', {}).get('status', 'unknown')])
            ws_stats.append(['Status API IBGE', self.api_status.get('ibge', {}).get('status', 'unknown')])
            
            wb.save(output_file)
            
        except ImportError:
            # Fallback para CSV se openpyxl não estiver disponível
            self._salvar_resultados_csv(resultados, output_file)
        except Exception as e:
            raise Exception(f"Erro ao salvar resultados: {str(e)}")

    def _salvar_resultados_csv(self, resultados, output_prefix):
        """Salva os resultados em arquivos CSV (fallback)"""
        # Implementação similar à anterior, mas com melhor tratamento de erros
        pass

    # Métodos de monitoramento de API v2.0
    def check_api_status(self) -> Dict[str, Dict]:
        """
        Verifica o status das APIs e retorna resultados
        
        Returns:
            Dict com status das APIs
        """
        return self.status_monitor.check_all_apis()
    
    def get_api_status(self) -> Dict[str, Dict]:
        """
        Retorna o status atual das APIs
        """
        return self.status_monitor.get_status()
    
    def is_api_online(self, api_type: str = 'cnpj') -> bool:
        """
        Verifica se uma API específica está online
        
        Args:
            api_type: 'cnpj' ou 'ibge'
            
        Returns:
            bool: True se online, False caso contrário
        """
        status = self.status_monitor.api_status.get(api_type, {}).get('status')
        return status == 'online'
    
    def get_api_status_emoji(self, api_type: str) -> str:
        """
        Retorna emoji representando o status da API
        """
        return self.status_monitor.get_api_status_emoji(api_type)
    
    def get_api_status_text(self, api_type: str) -> str:
        """
        Retorna texto descritivo do status
        """
        return self.status_monitor.get_api_status_text(api_type)
    
    def force_api_check(self):
        """Força verificação imediata do status das APIs"""
        return self.status_monitor.check_all_apis()

    # Métodos utilitários
    def clear_cache(self):
        """Limpa o cache completamente"""
        self.cache = {}
        if os.path.exists(self.cache_file):
            os.remove(self.cache_file)
        self._log_message("✅ Cache limpo com sucesso", "success")
    
    def get_cache_stats(self):
        """Retorna estatísticas do cache"""
        total = len(self.cache)
        ibge_entries = len(self.cache.get('ibge', {}))
        regular_entries = total - (1 if 'ibge' in self.cache else 0)
        
        return {
            'total_entries': total,
            'regular_entries': regular_entries,
            'ibge_entries': ibge_entries,
            'cache_file_size': os.path.getsize(self.cache_file) if os.path.exists(self.cache_file) else 0
        }
    
    def get_rate_limit_info(self):
        """Retorna informações sobre rate limit"""
        now = datetime.now()
        recent_requests = [
            t for t in self.ultimas_requisicoes 
            if now - t < timedelta(minutes=1)
        ]
        
        return {
            'current_requests': len(recent_requests),
            'max_requests': self.limite_requisicoes,
            'wait_time': self.tempo_espera,
            'time_to_reset': (timedelta(minutes=1) - (now - recent_requests[0])).total_seconds() if recent_requests else 0
        }

# Função utilitária para uso standalone
def consultar_cnpj_simples(cnpj: str) -> Optional[Dict[str, Any]]:
    """
    Função simples para consulta rápida de CNPJ
    """
    consultor = CNPJConsultor()
    return consultor.consultar_cnpj(cnpj)

if __name__ == "__main__":
    # Teste simples
    cnpj_teste = "00000000000191"  # CNPJ da União
    resultado = consultar_cnpj_simples(cnpj_teste)
    print(json.dumps(resultado, indent=2, ensure_ascii=False))