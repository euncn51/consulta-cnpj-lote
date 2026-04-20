"""
Módulo principal para consulta de CNPJs com correção na seleção de IE Ativa

Contém a classe CNPJConsultor que encapsula toda a lógica de:
- Comunicação com a API CNPJ WS
- Cache de resultados
- Processamento de dados
- Integração com IBGE para códigos municipais
- Seleção correta da IE Ativa
"""

from constants import API_CONFIG, EXCEL_STYLES

import os
import re
import json
import time
import unicodedata
from datetime import datetime, timedelta
from typing import Dict, Optional, List, Tuple, Any
from queue import Queue
import requests
import csv

class CNPJConsultor:
    """
    Classe principal para consulta e processamento de dados de CNPJ
    
    Atributos:
        api_url (str): URL base da API
        limite_requisicoes (int): Limite de requisições por minuto
        tempo_espera (int): Tempo de espera entre requisições (segundos)
        stop_processing (bool): Flag para parar processamento
        ultimas_requisicoes (list): Registro temporal das últimas requisições
        cache_file (str): Caminho do arquivo de cache
        cache (dict): Dados em cache
        session (requests.Session): Sessão HTTP persistente
    """
    
    def __init__(self):
        """Inicializa o consultor com configurações padrão"""
        self.api_url = API_CONFIG["base_url"]
        self.limite_requisicoes = API_CONFIG["default_rate_limit"]
        self.tempo_espera = API_CONFIG["default_wait_time"]
        self.stop_processing = False
        self.ultimas_requisicoes = []
        self.cache_file = API_CONFIG["cache_file"]
        self.cache = self._load_cache()
        
        self.session = requests.Session()
        self._configure_session()

    def _configure_session(self):
        """Configura a sessão HTTP com headers padrão para CNPJ WS"""
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'application/json',
            'Accept-Language': 'pt-BR,pt;q=0.9,en;q=0.8',
            'Referer': 'https://publica.cnpj.ws/',
            'Origin': 'https://publica.cnpj.ws'
        })
        self.session.timeout = API_CONFIG["timeout"]

    def _load_cache(self) -> Dict[str, Any]:
        """Carrega o cache do arquivo JSON"""
        try:
            if os.path.exists(self.cache_file):
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Erro ao carregar cache: {str(e)}")
        return {}

    def _save_cache(self) -> None:
        """Salva o cache atual no arquivo JSON"""
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Erro ao salvar cache: {str(e)}")

    @staticmethod
    def clean_text(text: str) -> str:
        """Normaliza e limpa texto removendo acentos e caracteres especiais"""
        if not isinstance(text, str):
            return str(text)
            
        text = unicodedata.normalize('NFD', text)
        text = text.encode('ascii', 'ignore').decode('utf-8')
        text = re.sub(r'[^a-zA-Z0-9\s]', ' ', text)
        return re.sub(r'\s+', ' ', text).strip()

    @staticmethod
    def clean_cnpj(cnpj: str) -> str:
        """Remove caracteres não numéricos do CNPJ"""
        return ''.join(filter(str.isdigit, str(cnpj)))

    def get_codigo_ibge(self, municipio: str, uf: str) -> str:
        """Obtém código IBGE para um município usando API do IBGE"""
        cache_key = f"{uf}_{municipio}"
        if cache_key in self.cache.get('ibge', {}):
            return self.cache['ibge'][cache_key]
            
        try:
            response = requests.get(
                f"https://servicodados.ibge.gov.br/api/v1/localidades/estados/{uf}/municipios",
                timeout=API_CONFIG["ibge_timeout"]
            )
            response.raise_for_status()
            
            for m in response.json():
                if m['nome'].upper() == municipio.upper():
                    if 'ibge' not in self.cache:
                        self.cache['ibge'] = {}
                    self.cache['ibge'][cache_key] = str(m['id'])
                    self._save_cache()
                    return str(m['id'])
        except Exception as e:
            print(f"Erro ao consultar IBGE: {str(e)}")
            
        return ""

    def _wait_if_needed(self) -> None:
        """Controla o rate limit esperando se necessário"""
        now = datetime.now()
        self.ultimas_requisicoes = [
            t for t in self.ultimas_requisicoes 
            if now - t < timedelta(minutes=1)
        ]
        
        if len(self.ultimas_requisicoes) >= self.limite_requisicoes:
            oldest = self.ultimas_requisicoes[0]
            wait_time = (timedelta(minutes=1) - (now - oldest)).total_seconds() + 1
            print(f"Rate limit atingido. Aguardando {wait_time:.1f} segundos...")
            time.sleep(wait_time)
            self.ultimas_requisicoes = []

    def consultar_cnpj(self, cnpj: str) -> Optional[Dict[str, Any]]:
        """Consulta um CNPJ na API CNPJ WS"""
        if self.stop_processing:
            return None

        cnpj_limpo = self.clean_cnpj(cnpj)

        # Verifica cache
        if cnpj_limpo in self.cache:
            cached_data = self.cache[cnpj_limpo]
            # Verifica se não é um erro em cache
            if not isinstance(cached_data, dict) or cached_data.get('ERRO'):
                return None
            return cached_data

        try:
            self._wait_if_needed()
            self.ultimas_requisicoes.append(datetime.now())

            print(f"Consultando CNPJ: {cnpj_limpo}")
            response = self.session.get(
                f"{self.api_url}{cnpj_limpo}", 
                timeout=API_CONFIG["timeout"]
            )

            if response.status_code == 200:
                dados = response.json()
                self.cache[cnpj_limpo] = dados
                self._save_cache()
                print(f"CNPJ {cnpj_limpo} consultado com sucesso")
                return dados

            elif response.status_code == 429:
                print(f"Rate limit atingido para CNPJ {cnpj_limpo}. Aguardando...")
                retry_after = int(response.headers.get('Retry-After', self.tempo_espera))
                time.sleep(retry_after + 2)
                return self.consultar_cnpj(cnpj)

            elif response.status_code == 404:
                print(f"CNPJ {cnpj_limpo} não encontrado.")
                self.cache[cnpj_limpo] = {'ERRO': 'CNPJ não encontrado'}
                self._save_cache()
                return None

            elif response.status_code == 400:
                print(f"CNPJ {cnpj_limpo} inválido.")
                self.cache[cnpj_limpo] = {'ERRO': 'CNPJ inválido'}
                self._save_cache()
                return None

            else:
                print(f"Erro na API: {response.status_code} - {response.text}")
                self.cache[cnpj_limpo] = {'ERRO': f'HTTP {response.status_code}'}
                self._save_cache()
                return None

        except requests.exceptions.Timeout:
            print(f"Timeout na consulta do CNPJ {cnpj_limpo}")
            self.cache[cnpj_limpo] = {'ERRO': 'Timeout'}
            self._save_cache()
            return None
        except requests.exceptions.RequestException as e:
            print(f"Erro na requisição: {str(e)}")
            self.cache[cnpj_limpo] = {'ERRO': f'Erro de conexão: {str(e)}'}
            self._save_cache()
            return None
        except Exception as e:
            print(f"Erro inesperado: {str(e)}")
            self.cache[cnpj_limpo] = {'ERRO': f'Erro inesperado: {str(e)}'}
            self._save_cache()
            return None

    def extrair_dados(self, dados: Dict[str, Any]) -> Dict[str, Any]:
        """Extrai e formata os dados relevantes da resposta da API CNPJ WS"""
        if not dados or 'estabelecimento' not in dados:
            return {'ERRO': 'Dados inválidos ou CNPJ não encontrado'}

        try:
            estabelecimento = dados.get('estabelecimento', {})
            
            # Processa inscrições estaduais
            ies = []
            inscricoes_estaduais = estabelecimento.get('inscricoes_estaduais', [])
            
            for insc in inscricoes_estaduais:
                ie = {
                    'NUMERO': insc.get('inscricao_estadual', ''),
                    'UF': insc.get('estado', {}).get('sigla', ''),
                    'TIPO': '',  # Não disponível na API CNPJ WS
                    'SITUACAO': 'Ativa' if insc.get('ativo', False) else 'Inativa',
                    'DATA_SITUACAO': insc.get('atualizado_em', '')[:10] if insc.get('atualizado_em') else '',
                    'ATIVA': 'Sim' if insc.get('ativo', False) else 'Não'
                }
                ies.append(ie)
            
            # Ordena por estado da empresa
            uf_empresa = estabelecimento.get('estado', {}).get('sigla', '')
            ies.sort(key=lambda x: (
                x.get('UF', '') != uf_empresa,  # False (0) primeiro para IEs do mesmo estado
            ))

            # Busca código IBGE
            municipio = estabelecimento.get('cidade', {}).get('nome', '')
            uf = estabelecimento.get('estado', {}).get('sigla', '')
            codigo_ibge = estabelecimento.get('cidade', {}).get('ibge_id', '')
            
            # Se não encontrou código IBGE, consulta API
            if not codigo_ibge and municipio and uf:
                codigo_ibge = self.get_codigo_ibge(municipio, uf)

            # Seleciona IE ativa - prioriza mesmo estado
            ie_ativa = None
            for ie in ies:
                if ie.get('UF') == uf and ie.get('ATIVA') == 'Sim':
                    ie_ativa = ie
                    break
            
            # Se não encontrou do mesmo estado, pega a primeira IE ativa
            if ie_ativa is None:
                for ie in ies:
                    if ie.get('ATIVA') == 'Sim':
                        ie_ativa = ie
                        break
            
            # Se ainda não encontrou, pega a primeira IE
            if ie_ativa is None and ies:
                ie_ativa = ies[0]

            # Endereço completo
            tipo_logradouro = estabelecimento.get('tipo_logradouro', '')
            logradouro = estabelecimento.get('logradouro', '')
            endereco_completo = f"{tipo_logradouro} {logradouro}".strip()

            # Telefones
            ddd1 = estabelecimento.get('ddd1', '')
            telefone1 = estabelecimento.get('telefone1', '')
            telefone = f"({ddd1}) {telefone1}" if ddd1 and telefone1 else telefone1

            # Data de situação cadastral formatada
            data_situacao = estabelecimento.get('data_situacao_cadastral', '')
            if data_situacao:
                try:
                    # Converte de YYYY-MM-DD para DD/MM/YYYY
                    data_obj = datetime.strptime(data_situacao, '%Y-%m-%d')
                    data_situacao_formatada = data_obj.strftime('%d/%m/%Y')
                except:
                    data_situacao_formatada = data_situacao
            else:
                data_situacao_formatada = ''

            # CORREÇÃO: Natureza jurídica - busca da estrutura correta
            natureza_juridica_descricao = ''
            if 'natureza_juridica' in dados:
                # Busca da raiz do objeto (estrutura atual da API)
                natureza_juridica = dados.get('natureza_juridica', {})
                natureza_juridica_descricao = natureza_juridica.get('descricao', '')
            elif 'empresa' in dados and 'natureza_juridica' in dados['empresa']:
                # Busca do objeto empresa (estrutura alternativa)
                natureza_juridica = dados['empresa'].get('natureza_juridica', {})
                natureza_juridica_descricao = natureza_juridica.get('descricao', '')

            # Cria dicionário com mapeamento correto
            resultado = {
                'CNPJ': estabelecimento.get('cnpj', ''),
                'RAZAO_SOCIAL': self.clean_text(dados.get('razao_social', '')),
                'NOME_FANTASIA': self.clean_text(estabelecimento.get('nome_fantasia', '') or ''),
                'NATUREZA_JURIDICA': self.clean_text(natureza_juridica_descricao),
                'ENDERECO': self.clean_text(endereco_completo),
                'NUMERO': estabelecimento.get('numero', ''),
                'COMPLEMENTO': self.clean_text(estabelecimento.get('complemento', '')),
                'BAIRRO': self.clean_text(estabelecimento.get('bairro', '')),
                'CEP': estabelecimento.get('cep', ''),
                'MUNICIPIO': municipio,
                'COD_IBGE': str(codigo_ibge) if codigo_ibge else '',
                'UF': uf,
                'TELEFONE': telefone,
                'EMAIL': estabelecimento.get('email', ''),
                'SITUACAO': estabelecimento.get('situacao_cadastral', ''),
                'DATA_SITUACAO_CADASTRAL': data_situacao_formatada,
                'DATA_CONSULTA': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'IE_ATIVA_NUMERO': ie_ativa.get('NUMERO', '') if ie_ativa else '',
                'IE_ATIVA_UF': ie_ativa.get('UF', '') if ie_ativa else '',
                'IE_ATIVA_SITUACAO': ie_ativa.get('SITUACAO', '') if ie_ativa else '',
                'IE_ATIVA_DATA': ie_ativa.get('DATA_SITUACAO', '') if ie_ativa else '',
                'INSCRICOES_ESTADUAIS': json.dumps(ies, ensure_ascii=False) if ies else '[]'
            }

            return resultado

        except Exception as e:
            print(f"Erro ao extrair dados: {str(e)}")
            return {'ERRO': f'Erro no processamento dos dados: {str(e)}'}

    def processar_planilha(self, input_file: str, output_file: str, 
                          coluna_cnpj: str, progress_queue: Queue) -> Tuple[bool, str]:
        """Processa um arquivo com múltiplos CNPJs"""
        try:
            self.stop_processing = False
            self.ultimas_requisicoes = []

            df = self._ler_arquivo(input_file)
            if not df or coluna_cnpj not in df[0]:
                progress_queue.put(("error", f"Coluna '{coluna_cnpj}' não encontrada"))
                return False, f"Coluna '{coluna_cnpj}' não encontrada"
            
            resultados = []
            total = len(df)
            
            for index, row in enumerate(df, 1):
                if self.stop_processing:
                    self._salvar_resultados(resultados, output_file, '_PARCIAL')
                    progress_queue.put(("info", "Processamento cancelado pelo usuário"))
                    return False, "Processamento cancelado"
                
                cnpj = row.get(coluna_cnpj, "")
                cnpj_limpo = self.clean_cnpj(str(cnpj))
                progress_queue.put(("progress", (index, total, cnpj_limpo)))
                
                if len(cnpj_limpo) != 14:
                    resultados.append({'CNPJ': cnpj, 'ERRO': 'CNPJ inválido'})
                    continue
                
                dados = self.consultar_cnpj(cnpj_limpo)
                if dados and 'estabelecimento' in dados:
                    resultados.append(self.extrair_dados(dados))
                else:
                    resultados.append({'CNPJ': cnpj_limpo, 'ERRO': 'Falha na consulta'})
            
            self._salvar_resultados(resultados, output_file)
            progress_queue.put(("success", f"Processamento concluído! Salvo em {output_file}"))
            return True, "Processamento concluído"
            
        except Exception as e:
            progress_queue.put(("error", f"Erro crítico: {str(e)}"))
            return False, str(e)

    def _ler_arquivo(self, input_file: str):
        """Lê arquivo CSV ou Excel"""
        ext = os.path.splitext(input_file)[1].lower()
        if ext == '.csv':
            with open(input_file, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                return list(reader)
        elif ext in ['.xlsx', '.xls']:
            from openpyxl import load_workbook
            wb = load_workbook(input_file, read_only=True)
            ws = wb.active
            data = []
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))
            return data
        raise ValueError("Formato de arquivo não suportado. Use Excel ou CSV.")

    def _salvar_resultados(self, resultados, output_file, suffix=''):
        """Salva resultados em arquivo Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        if suffix:
            output_file = output_file.replace('.xlsx', f'{suffix}.xlsx')
        
        # Prepara dados para as abas
        dados_ativas = []
        dados_historico = []
        
        for item in resultados:
            ies = []
            if item.get('INSCRICOES_ESTADUAIS') and item['INSCRICOES_ESTADUAIS'] != '[]':
                try:
                    ies = json.loads(item['INSCRICOES_ESTADUAIS'])
                    uf_empresa = item.get('UF', '')
                    ies.sort(key=lambda x: (x.get('UF', '') != uf_empresa))
                except json.JSONDecodeError:
                    pass
            
            # Adiciona ao histórico de IEs
            for ie in ies:
                dados_historico.append({
                    'CNPJ': item.get('CNPJ', ''),
                    'RAZAO_SOCIAL': item.get('RAZAO_SOCIAL', ''),
                    'IE_NUMERO': ie.get('NUMERO', ''),
                    'IE_UF': ie.get('UF', ''),
                    'IE_TIPO': ie.get('TIPO', ''),
                    'IE_SITUACAO': ie.get('SITUACAO', ''),
                    'IE_DATA_SITUACAO': ie.get('DATA_SITUACAO', ''),
                    'IE_ATIVA': ie.get('ATIVA', '')
                })
            
            # Seleciona IE ativa conforme a nova lógica
            ie_ativa = None
            for ie in ies:
                if ie.get('UF') == item.get('UF', ''):
                    ie_ativa = ie
                    break
            if ie_ativa is None and ies:
                ie_ativa = ies[0]
                
            row_ativa = {
                'CNPJ': item.get('CNPJ', ''),
                'RAZAO_SOCIAL': item.get('RAZAO_SOCIAL', ''),
                'NOME_FANTASIA': item.get('NOME_FANTASIA', ''),
                'NATUREZA_JURIDICA': item.get('NATUREZA_JURIDICA', ''),
                'ENDERECO': item.get('ENDERECO', ''),
                'NUMERO': item.get('NUMERO', ''),
                'COMPLEMENTO': item.get('COMPLEMENTO', ''),
                'BAIRRO': item.get('BAIRRO', ''),
                'CEP': item.get('CEP', ''),
                'MUNICIPIO': item.get('MUNICIPIO', ''),
                'COD_IBGE': item.get('COD_IBGE', ''),
                'UF': item.get('UF', ''),
                'TELEFONE': item.get('TELEFONE', ''),
                'EMAIL': item.get('EMAIL', ''),
                'SITUACAO': item.get('SITUACAO', ''),
                'DATA_SITUACAO_CADASTRAL': item.get('DATA_SITUACAO_CADASTRAL', ''),
                'DATA_CONSULTA': item.get('DATA_CONSULTA', ''),
                'IE_ATIVA_NUMERO': ie_ativa.get('NUMERO', '') if ie_ativa else '',
                'IE_ATIVA_UF': ie_ativa.get('UF', '') if ie_ativa else '',
                'IE_ATIVA_SITUACAO': ie_ativa.get('SITUACAO', '') if ie_ativa else '',
                'IE_ATIVA_DATA': ie_ativa.get('DATA_SITUACAO', '') if ie_ativa else ''
            }
            dados_ativas.append(row_ativa)
        
        # Cria workbook e aba principal
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = 'CNPJs'
        
        if dados_ativas:
            ws_main.append(list(dados_ativas[0].keys()))
            for row in dados_ativas:
                ws_main.append(list(row.values()))
        
        # Formatação da aba principal
        header_font = Font(**EXCEL_STYLES["header_font"])
        header_fill = PatternFill(**EXCEL_STYLES["header_fill"])
        
        for cell in ws_main[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = Border(bottom=Side(style='thin'))
        
        for col, width in EXCEL_STYLES["column_widths"]["main"].items():
            ws_main.column_dimensions[col].width = width
        
        # Aba de histórico
        if dados_historico:
            ws_hist = wb.create_sheet('Histórico IEs')
            ws_hist.append(list(dados_historico[0].keys()))
            for row in dados_historico:
                ws_hist.append(list(row.values()))
            
            for cell in ws_hist[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = Border(bottom=Side(style='thin'))
            
            for col, width in EXCEL_STYLES["column_widths"]["history"].items():
                ws_hist.column_dimensions[col].width = width
        
        wb.save(output_file)

    def _salvar_resultados_csv(self, resultados, output_prefix):
        """Salva os resultados em arquivos CSV"""
        # Aba principal
        campos_principal = [
            'CNPJ', 'RAZAO_SOCIAL', 'NOME_FANTASIA', 'NATUREZA_JURIDICA', 'ENDERECO', 'NUMERO', 'COMPLEMENTO', 'BAIRRO', 'CEP',
            'MUNICIPIO', 'COD_IBGE', 'UF', 'TELEFONE', 'EMAIL', 'SITUACAO', 'DATA_SITUACAO_CADASTRAL', 'DATA_CONSULTA',
            'IE_ATIVA_NUMERO', 'IE_ATIVA_UF', 'IE_ATIVA_SITUACAO', 'IE_ATIVA_DATA'
        ]
        # Aba histórico
        campos_hist = [
            'CNPJ', 'RAZAO_SOCIAL', 'IE_NUMERO', 'IE_UF', 'IE_TIPO', 'IE_SITUACAO', 'IE_DATA_SITUACAO', 'IE_ATIVA'
        ]

        dados_ativas = []
        dados_historico = []

        for item in resultados:
            ies = []
            if item.get('INSCRICOES_ESTADUAIS') and item['INSCRICOES_ESTADUAIS'] != '[]':
                try:
                    ies = json.loads(item['INSCRICOES_ESTADUAIS'])
                    uf_empresa = item.get('UF', '')
                    ies.sort(key=lambda x: (x.get('UF', '') != uf_empresa))
                except json.JSONDecodeError:
                    pass
            for ie in ies:
                dados_historico.append({
                    'CNPJ': item.get('CNPJ', ''),
                    'RAZAO_SOCIAL': item.get('RAZAO_SOCIAL', ''),
                    'IE_NUMERO': ie.get('NUMERO', ''),
                    'IE_UF': ie.get('UF', ''),
                    'IE_TIPO': ie.get('TIPO', ''),
                    'IE_SITUACAO': ie.get('SITUACAO', ''),
                    'IE_DATA_SITUACAO': ie.get('DATA_SITUACAO', ''),
                    'IE_ATIVA': ie.get('ATIVA', '')
                })
            # Seleciona IE ativa conforme a nova lógica
            ie_ativa = None
            for ie in ies:
                if ie.get('UF') == item.get('UF', ''):
                    ie_ativa = ie
                    break
            if ie_ativa is None and ies:
                ie_ativa = ies[0]
                
            row_ativa = {
                'CNPJ': item.get('CNPJ', ''),
                'RAZAO_SOCIAL': item.get('RAZAO_SOCIAL', ''),
                'NOME_FANTASIA': item.get('NOME_FANTASIA', ''),
                'NATUREZA_JURIDICA': item.get('NATUREZA_JURIDICA', ''),
                'ENDERECO': item.get('ENDERECO', ''),
                'NUMERO': item.get('NUMERO', ''),
                'COMPLEMENTO': item.get('COMPLEMENTO', ''),
                'BAIRRO': item.get('BAIRRO', ''),
                'CEP': item.get('CEP', ''),
                'MUNICIPIO': item.get('MUNICIPIO', ''),
                'COD_IBGE': item.get('COD_IBGE', ''),
                'UF': item.get('UF', ''),
                'TELEFONE': item.get('TELEFONE', ''),
                'EMAIL': item.get('EMAIL', ''),
                'SITUACAO': item.get('SITUACAO', ''),
                'DATA_SITUACAO_CADASTRAL': item.get('DATA_SITUACAO_CADASTRAL', ''),
                'DATA_CONSULTA': item.get('DATA_CONSULTA', ''),
                'IE_ATIVA_NUMERO': ie_ativa.get('NUMERO', '') if ie_ativa else '',
                'IE_ATIVA_UF': ie_ativa.get('UF', '') if ie_ativa else '',
                'IE_ATIVA_SITUACAO': ie_ativa.get('SITUACAO', '') if ie_ativa else '',
                'IE_ATIVA_DATA': ie_ativa.get('DATA_SITUACAO', '') if ie_ativa else ''
            }
            dados_ativas.append(row_ativa)

        # Salva aba principal
        with open(f"{output_prefix}_CNPJs.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=campos_principal)
            writer.writeheader()
            for row in dados_ativas:
                writer.writerow(row)

        # Salva aba histórico
        if dados_historico:
            with open(f"{output_prefix}_Historico_IEs.csv", "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=campos_hist)
                writer.writeheader()
                for row in dados_historico:
                    writer.writerow(row)